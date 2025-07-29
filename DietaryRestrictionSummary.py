import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

# Hide the root window
root = Tk()
root.withdraw()

# Ask the user to select an Excel file
file_path = askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])

# Define the mapping for dietary restriction codes
dietary_mapping = {
    "0": "No Restrictions",
    "1": "Vegetarian",
    "2": "Vegan",
    "3": "Gluten Free",
    "4": "No Dairy/Lactose Free",
    "5": "Halal",
    "6": "No Red Meat",
    "7": "No Pork",
    "8": "Kosher",
    "9": "Severe Allergy"
}

# Decode dietary restriction codes
def decode_meal_prefs(codes):
    try:
        codes_list = [code.strip() for code in str(codes).split(',')]
        decoded = [
            dietary_mapping[code] if code in dietary_mapping else f"Invalid Code: {code}"
            for code in codes_list
        ]
        return ', '.join(sorted(decoded))
    except Exception:
        return "Invalid Entry"

# Load the Excel file
if not file_path:
    print("No file selected. Exiting.")
    exit()

df = pd.read_excel(file_path, engine='openpyxl')

# Delete specified columns by index (0-based)
columns_to_delete = ["student_name_lowc", "session_id", "student_number", "begin_date", "perm_phone_num", "local_phone_num", "Expr1", "e_mail_ucs",
                     "e_mail_other" ]
df.drop(columns=columns_to_delete, inplace=True)

# Overwrite the original Excel file with the updated DataFrame (columns deleted)
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name='Simple Data', index=False)

# Make the data in 'Sheet1' into a table

wb = openpyxl.load_workbook(file_path)
ws = wb['Simple Data']

# Determine the range for the table
max_row = ws.max_row
max_col = ws.max_column
table_ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

table = Table(displayName="DataTable", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

# Autofit column widths
for col in ws.columns:
    max_length = 0
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    for cell in col:
        try:
            cell_value = str(cell.value) if cell.value is not None else ""
            if len(cell_value) > max_length:
                max_length = len(cell_value)
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

# Save the workbook with the table
wb.save(file_path)

# Decode the 'meal_prefs' column
df['meal_prefs'] = df['meal_prefs'].apply(decode_meal_prefs)

# Remove redundant dietary restrictions
def condense_restrictions(restriction_str):
    removal_mapping = {
        "Halal": {"No Pork"},
        "Kosher": {"No Pork"},
        "Vegetarian": {"No Red Meat", "No Pork"},
        "Vegan": {"Vegetarian", "No Dairy/Lactose Free"}
    }
    restrictions_set = set([r.strip() for r in restriction_str.split(',')])
    for key, values_to_remove in removal_mapping.items():
        if key in restrictions_set:
            restrictions_set -= values_to_remove
    # Remove duplicates and sort for consistency
    return ', '.join(sorted(restrictions_set))
df['meal_prefs'] = df['meal_prefs'].apply(condense_restrictions)

# Create a summary
summary = df.groupby('meal_prefs').size().reset_index(name='Count')
summary = summary.rename(columns={'meal_prefs': 'Dietary Restrictions'})

summary['Severe Allergy Comments'] = summary['Dietary Restrictions'].apply(
    lambda restriction: "; ".join(
        df.loc[
            (df['meal_prefs'] == restriction) &
            (df['meal_prefs'].str.contains("Severe Allergy", na=False)) &
            (df['comments'].notna()) &
            (df['comments'].str.strip() != "")
        ]['comments'].astype(str).str.strip().tolist()
    ) if "Severe Allergy" in restriction else ""
)

# Write the summary to a new sheet in the same Excel file
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    summary.to_excel(writer, sheet_name='Summary', index=False)

print("Summary sheet created successfully.")
# Notify the user that the process is complete
print("Dietary preferences have been decoded and a summary sheet has been created in the Excel file.")